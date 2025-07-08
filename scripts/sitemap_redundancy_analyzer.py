#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
快速 Sitemap 冗餘分析工具

專為快速、完整地分析 sitemap URL 冗餘情況而設計，優化了資料庫查詢和結果呈現。

核心功能：
1. 智能 Sitemap 發現（可選，或自動檢測）。
2. 高效獲取 Sitemap 中的所有 URL（支持 sitemap 索引，並發處理）。
3. 一次性查詢資料庫中所有相關頁面，避免多次查詢。
4. 清晰的數據覆蓋情況和冗餘分析報告。

使用方法：
# 自動發現 sitemap 並進行分析（預設輸出 Excel 到 data/ 資料夾，查詢全部時間）
poetry run python scripts/sitemap_redundancy_analyzer.py --site-id 14

# 手動指定 sitemap URL
poetry run python scripts/sitemap_redundancy_analyzer.py --site-id 14 \
    --sitemap-url "https://example.com/sitemap.xml" --days 30

# 指定自訂輸出路徑（Excel格式）
poetry run python scripts/sitemap_redundancy_analyzer.py --site-id 14 \
    --output-csv "reports/analysis.xlsx" --days 30

# 指定CSV格式（僅冗餘URL）
poetry run python scripts/sitemap_redundancy_analyzer.py --site-id 14 \
    --output-csv "reports/redundant_urls.csv" --days 30
"""

import argparse
import csv
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import quote, urljoin

import pandas as pd
import requests
from lxml import etree
from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, track
from rich.table import Table

# 將專案根目錄添加到 sys.path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# ruff: noqa: E402
from src.containers import Container
from src.services.database import Database

console = Console()

SITEMAP_NS = {"ns": "http://www.sitemaps.org/schemas/sitemap/0.9"}
COMMON_SITEMAP_PATHS = ["/sitemap.xml", "/sitemap_index.xml", "/sitemaps.xml"]
REQUEST_TIMEOUT_SHORT = 5
REQUEST_TIMEOUT_LONG = 30
MAX_WORKERS = 10  # 並發下載的線程數


class SitemapAnalyzer:
    """快速冗餘分析器"""

    def __init__(self, db: Database):
        self.db = db
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (compatible; SitemapAnalyzer/1.0; +https://github.com/user/repo)"
            }
        )

    def discover_sitemaps(self, domain: str) -> List[str]:
        """從 robots.txt 和常見路徑發現所有有效的 sitemaps"""
        if not domain.startswith(("http://", "https://")):
            domain = f"https://{domain}"

        console.print(Panel.fit(f"[bold blue]智能 Sitemap 發現[/bold blue]\n目標域名: {domain}"))

        # 1. 從 robots.txt 發現
        robots_url = urljoin(domain, "/robots.txt")
        sitemaps = []
        try:
            response = self.session.get(robots_url, timeout=REQUEST_TIMEOUT_SHORT)
            if response.status_code == 200:
                sitemap_pattern = re.compile(r"sitemap:\s*(.+)", re.IGNORECASE)
                sitemaps.extend([match.strip() for match in sitemap_pattern.findall(response.text)])
        except requests.RequestException as e:
            console.log(f"[yellow]無法讀取 {robots_url}: {e}[/yellow]")

        # 2. 檢查常見路徑
        for path in COMMON_SITEMAP_PATHS:
            sitemaps.append(urljoin(domain, path))

        # 3. 驗證所有 sitemaps
        validated_sitemaps = []
        for url in track(set(sitemaps), description="驗證 Sitemap..."):
            try:
                response = self.session.head(
                    url, timeout=REQUEST_TIMEOUT_SHORT, allow_redirects=True
                )
                if (
                    response.status_code == 200
                    and "xml" in response.headers.get("content-type", "").lower()
                ):
                    validated_sitemaps.append(str(response.url))  # 使用重定向後的最終 URL
                    console.print(f"✅ 找到有效 Sitemap: {response.url}")
            except requests.RequestException:
                continue

        if not validated_sitemaps:
            console.print("[bold red]❌ 未找到任何有效的 sitemap[/bold red]")
            return []

        console.print(
            f"[bold green]發現 {len(validated_sitemaps)} 個有效的 Sitemaps，將全部使用[/bold green]"
        )
        return validated_sitemaps

    def discover_and_select_sitemap(self, domain: str, auto_select: bool = True) -> Optional[str]:
        """保持向後兼容性的方法，返回第一個發現的 sitemap"""
        all_sitemaps = self.discover_sitemaps(domain)
        return all_sitemaps[0] if all_sitemaps else None

    def _fetch_and_parse_single_sitemap(self, url: str) -> List[str]:
        """獲取並解析單個 sitemap 或 sitemap 索引，返回 URL 列表"""
        try:
            response = self.session.get(url, timeout=REQUEST_TIMEOUT_LONG)
            response.raise_for_status()  # 拋出 HTTP 錯誤
            root = etree.fromstring(response.content, parser=etree.XMLParser(recover=True))

            # 檢查是 sitemap 索引還是 url 集合
            if root.tag.endswith("sitemapindex"):
                xpath_expr = "//ns:sitemap/ns:loc"
            elif root.tag.endswith("urlset"):
                xpath_expr = "//ns:url/ns:loc"
            else:
                # 可能是沒有 namespace 的 sitemap
                if root.tag == "sitemapindex":
                    xpath_expr = "//sitemap/loc"
                else:
                    xpath_expr = "//url/loc"

            # 提取文本並過濾掉 None 或空字串
            return [
                elem.text for elem in root.xpath(xpath_expr, namespaces=SITEMAP_NS) if elem.text
            ]

        except requests.HTTPError as e:
            console.log(f"[red]HTTP 錯誤 {e.response.status_code} 於 {url}[/red]")
        except Exception as e:
            console.log(f"[red]處理 {url} 失敗: {e}[/red]")
        return []

    def fetch_sitemap_urls(self, sitemap_url: str) -> List[str]:
        """高效獲取並解析 Sitemap，提取所有 URL，支持並發處理索引"""
        console.print(f"\n🔍 正在獲取 Sitemap: {sitemap_url}")

        initial_urls = self._fetch_and_parse_single_sitemap(sitemap_url)
        if not initial_urls:
            return []

        # 判斷是 sitemap 索引還是普通的 sitemap
        if initial_urls[0].endswith(".xml"):
            console.print(
                f"📄 檢測到 Sitemap 索引，包含 {len(initial_urls)} 個子 sitemap，開始並發處理..."
            )
            all_urls = []
            with Progress(console=console) as progress:
                task = progress.add_task("[cyan]解析子 Sitemap...", total=len(initial_urls))
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    future_to_url = {
                        executor.submit(self._fetch_and_parse_single_sitemap, url): url
                        for url in initial_urls
                    }
                    for future in as_completed(future_to_url):
                        try:
                            urls_from_sub = future.result()
                            all_urls.extend(urls_from_sub)
                        except Exception as e:
                            url = future_to_url[future]
                            console.log(f"[red]獲取子 sitemap {url} 時發生嚴重錯誤: {e}[/red]")
                        progress.update(task, advance=1)
            urls = all_urls
        else:
            urls = initial_urls

        console.print("\n📊 [bold]Sitemap URL 提取結果[/bold]")
        console.print(f"   🎯 Sitemap 總 URL 數: {len(urls):,} 個")
        console.print(f"   📄 來源: {sitemap_url}")
        return urls

    def get_db_pages_and_coverage(self, site_id: int, days: Optional[int]) -> Tuple[Set[str], dict]:
        """一次性獲取資料庫中的不重複頁面及數據覆蓋情況"""

        console.print("\n🔍 正在查詢資料庫...")

        date_clause = ""
        time_range_text = "全部時間"

        if days:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            date_clause = "AND date BETWEEN ? AND ?"
            params = [site_id, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")]
            time_range_text = f"最近 {days} 天"
        else:
            params = [site_id]

        # 查詢數據覆蓋情況
        coverage_query = f"""
            SELECT date, COUNT(*) FROM gsc_performance_data
            WHERE site_id = ? {date_clause}
            GROUP BY date ORDER BY date DESC
        """
        with self.db._lock:
            coverage_results = self.db._connection.execute(coverage_query, params).fetchall()

        actual_days = len(coverage_results)
        total_records = sum(row[1] for row in coverage_results)

        console.print(f"📊 [bold]資料庫數據覆蓋情況[/bold] (查詢範圍: {time_range_text})")
        if days and days > 0:
            console.print(
                f"   ✅ 實際有數據天數: {actual_days} 天 ({actual_days / days * 100:.1f}%)"
            )
        else:
            console.print(f"   ✅ 實際有數據天數: {actual_days} 天")

        if actual_days > 0:
            console.print(
                f"   📅 數據日期範圍: {coverage_results[-1][0]} 到 {coverage_results[0][0]}"
            )
        console.print(f"   📈 總數據點記錄數: {total_records:,} 筆")

        # 查詢不重複頁面
        pages_query = (
            f"SELECT DISTINCT page FROM gsc_performance_data WHERE site_id = ? {date_clause}"
        )
        with self.db._lock:
            page_results = self.db._connection.execute(pages_query, params).fetchall()

        db_pages = {row[0] for row in page_results}
        console.print(f"   🎯 有數據的獨立頁面URL數: {len(db_pages):,} 個")

        coverage_info = {
            "queried_days": time_range_text,
            "actual_days": actual_days,
            "db_unique_pages": len(db_pages),
        }
        return db_pages, coverage_info

    def analyze_and_display(
        self,
        sitemap_urls: List[str],
        db_pages: Set[str],
        site_info: Dict,
        coverage_info: Dict,
        output_csv_path: Optional[str] = None,
        site_id: Optional[int] = None,
        days: Optional[int] = None,
    ):
        """分析冗餘並顯示結果"""
        console.print("\n🔍 正在進行冗餘分析...")

        sitemap_set = set(sitemap_urls)

        urls_in_db = sitemap_set.intersection(db_pages)
        urls_not_in_db = sitemap_set - db_pages

        # 創建結果面板
        title = f"[bold]Sitemap 冗餘分析報告 for {site_info.get('name', '未知')}[/bold]"
        table = Table(title=title, show_header=False, box=None, padding=(0, 2))
        table.add_column(style="cyan")
        table.add_column(style="magenta")

        table.add_row("\n[bold]Sitemap 統計[/bold]", "")
        table.add_row("Sitemap 總 URL 數", f"{len(sitemap_urls):,}")
        table.add_row("去重後獨立 URL 數", f"{len(sitemap_set):,}")

        table.add_row("\n[bold]資料庫統計[/bold]", f"({coverage_info['queried_days']})")
        table.add_row("有數據的獨立 URL 數", f"{coverage_info['db_unique_pages']:,}")

        table.add_row("\n[bold]冗餘分析 (Sitemap vs 資料庫)[/bold]", "")
        table.add_row("✅ 有數據的 Sitemap URL", f"{len(urls_in_db):,}")
        table.add_row("❌ 無數據的 Sitemap URL", f"{len(urls_not_in_db):,}")

        if len(sitemap_set) > 0:
            redundancy_rate = len(urls_not_in_db) / len(sitemap_set) * 100
            coverage_rate = len(urls_in_db) / len(sitemap_set) * 100

            # 根據冗餘率顯示不同顏色的結果
            color = "green"
            if redundancy_rate > 50:
                color = "red"
            elif redundancy_rate > 20:
                color = "yellow"

            table.add_row(
                f"[{color}]冗餘率 (無數據的 URL 佔比)[/{color}]",
                f"[{color}]{redundancy_rate:.1f}%[/{color}]",
            )
            table.add_row(
                "[green]覆蓋率 (有數據的 URL 佔比)[/green]", f"[green]{coverage_rate:.1f}%[/green]"
            )

        console.print(Panel(table, expand=False))

        # 如果提供了輸出路徑，則將詳細分析結果寫入 Excel
        if output_csv_path:
            output_path = Path(output_csv_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            try:
                # 判斷輸出格式：如果檔案副檔名是 .xlsx，則輸出 Excel；否則輸出 CSV
                if output_path.suffix.lower() == ".xlsx":
                    if site_id is not None:
                        self._export_to_excel(
                            output_path,
                            sitemap_urls,
                            urls_in_db,
                            urls_not_in_db,
                            site_info,
                            coverage_info,
                            redundancy_rate if len(sitemap_set) > 0 else 0,
                            coverage_rate if len(sitemap_set) > 0 else 0,
                            site_id,
                            days,
                        )
                    else:
                        console.print(
                            "[bold red]錯誤：無法生成Excel檔案，缺少site_id參數[/bold red]"
                        )
                else:
                    # 保持原有的 CSV 格式（只有無數據的 URL）
                    with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(["URL"])  # 寫入標頭
                        for url in sorted(list(urls_not_in_db)):
                            writer.writerow([url])
                    console.print(
                        f"\n💾 [bold green]無數據的冗餘 URL 列表已儲存至: "
                        f"{output_path}[/bold green]"
                    )
            except Exception as e:
                console.print(f"\n[bold red]錯誤：無法寫入檔案 {output_path}: {e}[/bold red]")

    def _export_to_excel(
        self,
        output_path: Path,
        sitemap_urls: List[str],
        urls_in_db: Set[str],
        urls_not_in_db: Set[str],
        site_info: Dict,
        coverage_info: Dict,
        redundancy_rate: float,
        coverage_rate: float,
        site_id: int,
        days: Optional[int],
    ):
        """將分析結果導出為 Excel 檔案，包含多個工作表"""

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # 工作表1：分析報告摘要
            summary_data = {
                "項目": [
                    "網站名稱",
                    "網站ID",
                    "Sitemap 總 URL 數",
                    "Sitemap 去重後獨立 URL 數",
                    "GSC performace 中的獨立 URL 數",
                    "✅ 擁有 GSC performace 的 URL（Sitemap）",
                    "❌ 沒有 GSC performace 數據的  URL（Sitemap）",
                    "冗餘率 (%)",
                    "覆蓋率 (%)",
                    "查詢時間範圍",
                    "實際有數據天數",
                ],
                "數值": [
                    site_info.get("name", "未知"),
                    site_info.get("id", "未知"),
                    len(sitemap_urls),
                    len(set(sitemap_urls)),
                    coverage_info["db_unique_pages"],
                    len(urls_in_db),
                    len(urls_not_in_db),
                    f"{redundancy_rate:.1f}%",
                    f"{coverage_rate:.1f}%",
                    coverage_info["queried_days"],
                    coverage_info["actual_days"],
                ],
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="分析報告", index=False)

            # 工作表2：有數據的URL列表
            if urls_in_db:
                # 對 URL 進行編碼
                encoded_urls = [
                    quote(url, safe=":/?#[]@!$&'()*+,;=") for url in sorted(list(urls_in_db))
                ]
                urls_with_data_df = pd.DataFrame(
                    {"URL": encoded_urls, "狀態": ["有數據"] * len(urls_in_db)}
                )
                urls_with_data_df.to_excel(
                    writer, sheet_name="有 GSC performace 的 URL", index=False
                )

            # 工作表3：無數據的URL列表（冗餘）
            if urls_not_in_db:
                # 對 URL 進行編碼
                encoded_urls = [
                    quote(url, safe=":/?#[]@!$&'()*+,;=") for url in sorted(list(urls_not_in_db))
                ]
                urls_without_data_df = pd.DataFrame(
                    {"URL": encoded_urls, "狀態": ["無數據"] * len(urls_not_in_db)}
                )
                urls_without_data_df.to_excel(
                    writer, sheet_name="無 GSC performace URL", index=False
                )

            # 工作表4：每月平均表現表
            if urls_in_db:
                monthly_performance = self._get_monthly_performance(site_id, list(urls_in_db), days)
                if monthly_performance:
                    monthly_df = pd.DataFrame(monthly_performance)
                    monthly_df.to_excel(writer, sheet_name="每月平均表現表", index=False)

                    # 設置 Excel 格式，讓關鍵字欄位支持換行顯示
                    worksheet = writer.sheets["每月平均表現表"]

                    # 找到關鍵字欄位的索引
                    if monthly_performance:
                        columns = list(monthly_performance[0].keys())
                        if "關鍵字" in columns:
                            keyword_col_idx = columns.index("關鍵字") + 1  # Excel 列索引從1開始

                            # 設置關鍵字欄位的格式
                            for row_idx in range(
                                2, len(monthly_performance) + 2
                            ):  # 從第2行開始（跳過標題行）
                                cell = worksheet.cell(row=row_idx, column=keyword_col_idx)
                                cell.alignment = cell.alignment.copy(wrapText=True)

                            # 設置欄位寬度
                            worksheet.column_dimensions[
                                chr(ord("A") + keyword_col_idx - 1)
                            ].width = 50

            # 應用樣式
            self._apply_excel_styles(writer)

        console.print(f"\n💾 [bold green]詳細分析報告已儲存至: {output_path}[/bold green]")
        console.print("📊 Excel 檔案包含以下工作表：")
        console.print("   • 分析報告 - 摘要統計")
        console.print("   • 有 GSC performace 的 URL - 在GSC資料庫中有數據的URL")
        console.print("   • 無 GSC performace URL - 在Sitemap中但GSC資料庫無數據的URL")
        console.print("   • 每月平均表現表 - 有數據URL的月度表現統計")

    def _apply_excel_styles(self, writer):
        """為 Excel 檔案應用樣式"""
        from openpyxl.styles import Alignment, Font, PatternFill

        # 定義樣式
        header_font = Font(bold=True, color="FFFFFF")  # 粗體白字
        header_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )  # 黑底

        highlight_font = Font(color="000000")  # 黑字
        highlight_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )  # 螢光筆背景（黃色）

        # 應用樣式到所有工作表
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            # 為 header 行添加樣式（第一行）
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 為分析報告中的特定行添加螢光筆樣式
            if sheet_name == "分析報告":
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    item_cell = row[0]  # 項目欄位
                    value_cell = row[1]  # 數值欄位

                    # 檢查是否為冗餘率或覆蓋率
                    if item_cell.value and (
                        "冗餘率" in str(item_cell.value) or "覆蓋率" in str(item_cell.value)
                    ):
                        item_cell.font = highlight_font
                        item_cell.fill = highlight_fill
                        value_cell.font = highlight_font
                        value_cell.fill = highlight_fill

            # 調整列寬
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大寬度限制為50
                worksheet.column_dimensions[column_letter].width = adjusted_width

    def _get_monthly_performance(
        self, site_id: int, urls_with_data: List[str], days: Optional[int]
    ) -> List[Dict]:
        """獲取有數據URL的每月平均表現"""
        console.print("\n📈 正在獲取每月平均表現數據...")

        # 構建日期篩選條件
        date_clause = ""
        params: List[str] = [str(site_id)]

        if days and days > 0:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            date_clause = "AND date >= ?"
            params.append(start_date.strftime("%Y-%m-%d"))

        # 創建URL參數佔位符
        url_placeholders = ",".join("?" * len(urls_with_data))
        params.extend(urls_with_data)

        # 查詢每月平均表現（修改為加總點擊數和曝光數，並新增關鍵字相關欄位）
        query = f"""
        SELECT
            page,
            strftime('%Y-%m', date) as month,
            SUM(clicks) as total_clicks,
            SUM(impressions) as total_impressions,
            AVG(ctr) as avg_ctr,
            AVG(position) as avg_position,
            COUNT(*) as record_count,
            REPLACE(GROUP_CONCAT(DISTINCT query), ',', CHAR(10)) as keywords,
            COUNT(DISTINCT query) as keyword_count
        FROM gsc_performance_data
        WHERE site_id = ? {date_clause}
        AND page IN ({url_placeholders})
        GROUP BY page, strftime('%Y-%m', date)
        ORDER BY page, month
        """

        try:
            with self.db._lock:
                results = self.db._connection.execute(query, params).fetchall()

            performance_data = []
            for row in results:
                # 對 URL 進行編碼
                encoded_url = quote(row[0], safe=":/?#[]@!$&'()*+,;=")

                performance_data.append(
                    {
                        "URL": encoded_url,
                        "月份": row[1],
                        "總點擊數": int(row[2] or 0),
                        "總曝光數": int(row[3] or 0),
                        "平均點擊率": round((row[4] or 0) * 100, 3),  # 轉換為百分比
                        "平均排名": round(row[5] or 0, 2),
                        "記錄數": row[6],
                        "關鍵字": row[7] or "",
                        "關鍵字數": row[8] or 0,
                    }
                )

            console.print(f"   ✅ 獲取 {len(performance_data)} 條月度表現記錄")
            return performance_data

        except Exception as e:
            console.print(f"[red]獲取月度表現數據時發生錯誤: {e}[/red]")
            return []


def main():
    """主執行函數"""
    parser = argparse.ArgumentParser(description="快速 Sitemap 冗餘分析工具")
    parser.add_argument("--site-id", type=int, required=True, help="要分析的網站 ID")
    parser.add_argument(
        "--sitemap-url", type=str, action="append", help="手動指定 Sitemap URL（可多次使用）"
    )
    parser.add_argument("--days", type=int, help="查詢天數範圍（可選，預設查詢全部時間）")
    parser.add_argument(
        "--interactive-discovery", action="store_true", help="強制進行交互式 Sitemap 選擇"
    )
    parser.add_argument(
        "--single-sitemap", action="store_true", help="只使用第一個發現的 sitemap（舊行為）"
    )
    parser.add_argument(
        "--output-csv",
        type=str,
        help="將分析結果導出到指定檔案 (.xlsx=Excel多工作表, .csv=僅冗餘URL, "
        "預設輸出Excel到data/資料夾)",
    )
    parser.add_argument(
        "--no-smart-discovery",
        action="store_true",
        help="暫停智能 Sitemap 發現功能，需要手動指定 --sitemap-url",
    )

    args = parser.parse_args()

    console.print(Panel.fit("[bold blue]快速 Sitemap 冗餘分析工具[/bold blue]"))

    try:
        container = Container()
        db = container.database()
        console.print("✅ 資料庫連接成功")
    except Exception as e:
        console.print(f"[bold red]錯誤：無法連接資料庫: {e}[/bold red]")
        sys.exit(1)

    site_info = db.get_site_by_id(args.site_id)
    if not site_info:
        console.print(f"[bold red]錯誤：找不到 ID 為 {args.site_id} 的站點[/bold red]")
        sys.exit(1)

    analyzer = SitemapAnalyzer(db)

    # 獲取 Sitemap URL
    sitemap_urls_to_fetch = args.sitemap_url
    if not sitemap_urls_to_fetch:
        if args.no_smart_discovery:
            console.print(
                "[bold yellow]智能 Sitemap 發現功能已暫停，請使用 --sitemap-url "
                "手動指定 Sitemap URL[/bold yellow]"
            )
            console.print("範例：--sitemap-url 'https://example.com/sitemap.xml'")
            sys.exit(1)

        domain = site_info["domain"].replace("sc-domain:", "")

        if args.single_sitemap:
            # 使用舊行為，只選擇一個 sitemap
            discovered_url = analyzer.discover_and_select_sitemap(
                domain, auto_select=not args.interactive_discovery
            )
            if discovered_url:
                sitemap_urls_to_fetch = [discovered_url]
            else:
                sys.exit(1)
        else:
            # 新行為：使用所有發現的 sitemaps
            discovered_urls = analyzer.discover_sitemaps(domain)
            if discovered_urls:
                sitemap_urls_to_fetch = discovered_urls
            else:
                sys.exit(1)

    # 統計總 URL 數
    all_sitemap_urls = []
    total_urls_from_all_sources = 0

    console.print(f"\n🔍 [bold]開始處理 {len(sitemap_urls_to_fetch)} 個 Sitemap 來源[/bold]")

    for i, url in enumerate(sitemap_urls_to_fetch, 1):
        console.print(
            f"\n📄 [bold cyan]處理 Sitemap {i}/{len(sitemap_urls_to_fetch)}:[/bold cyan] {url}"
        )
        urls_from_this_source = analyzer.fetch_sitemap_urls(url)
        if urls_from_this_source:
            all_sitemap_urls.extend(urls_from_this_source)
            total_urls_from_all_sources += len(urls_from_this_source)
            console.print(f"   ✅ 從此來源獲取: {len(urls_from_this_source):,} 個 URL")
        else:
            console.print("   ❌ 此來源無法獲取任何 URL")

    # 去重統計
    unique_urls = set(all_sitemap_urls)
    console.print("\n📊 [bold green]所有 Sitemap 合併統計:[/bold green]")
    console.print(f"   🎯 總 URL 數: {total_urls_from_all_sources:,} 個")
    console.print(f"   🔄 去重後: {len(unique_urls):,} 個")
    if total_urls_from_all_sources > 0:
        duplicate_rate = (
            (total_urls_from_all_sources - len(unique_urls)) / total_urls_from_all_sources * 100
        )
        console.print(f"   📉 重複率: {duplicate_rate:.1f}%")
    else:
        console.print("   📉 重複率: 0%")

    if not all_sitemap_urls:
        console.print("[bold yellow]未能從指定的 Sitemap 中提取任何 URL。[/bold yellow]")
        sys.exit(0)

    db_pages, coverage_info = analyzer.get_db_pages_and_coverage(args.site_id, args.days)

    # 如果沒有指定輸出路徑，則預設輸出到 data 資料夾（Excel格式）
    output_csv_path = args.output_csv
    if not output_csv_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        site_name = site_info.get("name", f"site_{args.site_id}").replace(" ", "_")
        filename = f"sitemap_redundancy_{site_name}_{timestamp}.xlsx"
        output_csv_path = f"data/{filename}"

    analyzer.analyze_and_display(
        all_sitemap_urls,
        db_pages,
        site_info,
        coverage_info,
        output_csv_path,
        args.site_id,
        args.days,
    )


if __name__ == "__main__":
    main()
